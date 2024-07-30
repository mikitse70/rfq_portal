from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from datetime import date
from io import BytesIO
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from itertools import groupby
from operator import attrgetter
from openpyxl import Workbook
from django.contrib.auth import update_session_auth_hash
from urllib import request
from django.contrib.auth.decorators import user_passes_test
from django.urls import reverse
import requests
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.http import Http404, HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render, redirect, HttpResponse
import pandas as pd
from django.utils.http import urlencode
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
from hashids import Hashids
from quotation.models import QuotationLine, QuotationRequest, SupplierBid
from django.contrib.auth import authenticate, login, logout
from cryptography.fernet import Fernet
from django.contrib.auth.models import User
from django.db.models import Q
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


hashids = Hashids(salt="your_unique_salt_here", min_length=8)


# function to send SMS message
def send_message(host, port, username, password, number, message):
    message = message.encode('utf-8')
    url = f"http://{host}:{port}/http/send-message/"
    params = {
        'username': username,
        'password': password,
        'to': number,
        'message': message,
    }
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        return response.text
    except requests.RequestException as e:
        return f"Error: {e}"

# view to select accounts with user role


def user_in_users_group(user):
    return user.groups.filter(name='user').exists()


# view to select accounts with sr role
def user_in_supplier_group(user):
    return user.groups.filter(name='supplier').exists()


# view to download specification(user)
@login_required(login_url='login')
def download_specification(request, quotation_id):
    quotation = get_object_or_404(QuotationRequest, pk=quotation_id)
    file_path = os.path.join(settings.MEDIA_ROOT, str(quotation.specification))
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(
                fh.read(), content_type="application/octet-stream")
            response['Content-Disposition'] = 'inline; filename=' + \
                os.path.basename(file_path)
            return response
    raise Http404

# view to download BOQ(user)


@login_required(login_url='login')
def download_boq(request, quotation_id):
    quotation = get_object_or_404(QuotationRequest, pk=quotation_id)
    file_path = os.path.join(settings.MEDIA_ROOT, str(quotation.boq))
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(
                fh.read(), content_type="application/octet-stream")
            response['Content-Disposition'] = 'inline; filename=' + \
                os.path.basename(file_path)
            return response
    raise Http404

# view to download attachment (supplier)


@login_required(login_url='login')
def download_attachement(request, supplier_bid):
    bid = get_object_or_404(SupplierBid, pk=supplier_bid)
    file_path = os.path.join(settings.MEDIA_ROOT, str(bid.file))
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(
                fh.read(), content_type="application/octet-stream")
            response['Content-Disposition'] = 'inline; filename=' + \
                os.path.basename(file_path)
            return response
    raise Http404

# view to download BOQ (supplier)


@login_required(login_url='login')
def download_supplier_boq(request, supplier_bid):
    bid = get_object_or_404(SupplierBid, pk=supplier_bid)
    file_path = os.path.join(settings.MEDIA_ROOT, str(bid.boq_total))
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(
                fh.read(), content_type="application/octet-stream")
            response['Content-Disposition'] = 'inline; filename=' + \
                os.path.basename(file_path)
            return response
    raise Http404


# logout view
def logout_view(request):
    request.session.flush()
    return redirect('login')

# login view


def login_view(request):
    id_value = request.GET.get('id')
    checked_lines_value = request.GET.get('checked_lines')
    request.session['id'] = id_value
    request.session['checked_lines'] = checked_lines_value
    request.session.modified = True
    urls = f'?id={id_value}&checked_lines={checked_lines_value}'
    if request.method == 'POST':
        username = request.POST.get('email')
        password = request.POST.get('password')
        user = None
        if username:
            user = authenticate(request, username=username, password=password)
        if not user:

            users_by_last_name = User.objects.filter(last_name=username)
            if users_by_last_name.exists():
                for user_by_last_name in users_by_last_name:
                    user = authenticate(
                        request, username=user_by_last_name.username, password=password)
                    if user:
                        break

        if not user:

            user = authenticate(request, email=username, password=password)

        if user is not None:
            login(request, user)
            groups = user.groups.all()
            if groups.filter(name='admin').exists():
                return redirect('admin_dashboard')
            elif groups.filter(name='supplier').exists():

                id_value = request.GET.get('id')
                checked_lines_value = request.GET.get('checked_lines')
                request.session['id'] = id_value
                request.session['checked_lines'] = checked_lines_value
                request.session.modified = True
                if request.session.get('id'):
                    return redirect('supplier_dashboard')
                else:
                    messages.error(
                        request, f'Use the link sent to you through emalil or SMS')
                    return redirect('login')

            elif groups.filter(name='user').exists():
                return redirect('user_dashboard')
            else:
                messages.error(request, f'Unable to login')
                redirect_url = reverse(
                    'login') + f'?id={id_value}&checked_lines={checked_lines_value}'
                return redirect(redirect_url)
        else:
            messages.error(request, f'Wrong password')
            redirect_url = reverse(
                'login') + f'?id={id_value}&checked_lines={checked_lines_value}'
            return redirect(redirect_url)
    else:
        context = {'urls': urls}
        return render(request, 'index.html', context)

# user dashboard view


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def userDashboard(request):
    current_date = date.today()
    quotations = QuotationRequest.objects.filter(
        deadline_date__gte=current_date)
    search_keyword = request.GET.get('search')
    if search_keyword:
        quotations = quotations.filter(
            quotation_number__icontains=search_keyword)

    paginator = Paginator(quotations, 10)
    page = request.GET.get('page')
    try:
        quotations = paginator.page(page)
    except PageNotAnInteger:
        quotations = paginator.page(1)
    except EmptyPage:
        quotations = paginator.page(paginator.num_pages)

    context = {'quotations': quotations}
    return render(request, 'user/index.html', context)

# view to request quotation


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def userquotation_request(request):
    if request.method == 'POST':

        if 'submit' in request.POST:
            try:
                item_name = request.POST.get('item_type')
                request_no = request.POST.get('request_no')
                quantity = request.POST.get('quantity')
                delivery_place = request.POST.get('delivery_place')
                description = request.POST.get('description')
                book_type = request.POST.get('book_type')
                deadline_date = request.POST.get('deadline_date')
                specification = request.FILES.get('specification')
                boq = request.FILES.get('boq')
                rfq_ref = request.POST.get('rfq_ref')
                if QuotationRequest.objects.filter(quotation_number=request_no).exists():
                    messages.error(request, 'Quotation number already exists')
                    return redirect('user_dashboard')

                quotation_request = QuotationRequest(
                    user_id=request.user,
                    quotation_number=request_no,
                    delivery_place=delivery_place,
                    specification=specification,
                    boq=boq,
                    deadline_date=deadline_date,
                    rfq_ref=rfq_ref
                )
                quotation_request.save()

                quotation_line = QuotationLine(
                    quotation_request=quotation_request,
                    item_description=description,
                    quantity=quantity,
                    book_type=book_type,
                    line_number=1
                )
                quotation_line.save()

                messages.success(request, 'request created successfully')
                return redirect('user_dashboard')

            except Exception as e:
                messages.error(request, f'Success message')
                return HttpResponse(f"Error: {e}")

        elif 'importexcel' in request.POST and request.FILES.get('excelfile'):
            try:
                excel_file = request.FILES['excelfile']
                df = pd.read_excel(excel_file)
                user = request.user
                if QuotationRequest.objects.filter(quotation_number=df.iloc[0]["Requisition"]).exists():
                    messages.error(request, 'Quotation number already exists')
                    return redirect('user_dashboard')

                quotation_request = QuotationRequest(
                    user_id=request.user,

                    quotation_number=df.iloc[0]["Requisition"],

                    rfq_ref=df.iloc[0]["RfqRef"],

                    deadline_date=pd.to_datetime(
                        df.iloc[0]["Deadline Date"], format="%d/%m/%Y"),
                )
                quotation_request.save()

                for index, row in df.iterrows():
                    quotation_line = QuotationLine(
                        quotation_request=quotation_request,
                        item_description=row["Stock code description"],
                        quantity=row["Quantity"],
                        book_type=row["Stock Type"],
                        line_number=row["Line"],
                    )
                    quotation_line.save()

                messages.success(request, 'Excell File Imported Successfuly')
                return redirect('user_dashboard')

            except Exception as e:
                messages.error(request, f"Error importing Excel data: {e}")
                return redirect('quotation_request')
    current_date = date.today()
    context = {'current_date': current_date}
    return render(request, 'user/quotation_request.html', context)

# user bids view(user)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def user_bids(request):
    bids = QuotationRequest.objects.filter(
        supplierbid__isnull=False).distinct()
    search_keyword = request.GET.get('search')
    if search_keyword:
        bids = bids.filter(quotation_number__icontains=search_keyword)

    paginator = Paginator(bids, 10)
    page = request.GET.get('page')
    try:
        bids = paginator.page(page)
    except PageNotAnInteger:
        bids = paginator.page(1)
    except EmptyPage:
        bids = paginator.page(paginator.num_pages)

    context = {'bids': bids}
    return render(request, 'user/bids.html', context)

# view edit request(user)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def edit_request(request,  quotation_request_id):
    quotation_request = QuotationRequest.objects.get(id=quotation_request_id)
    quotation_request = get_object_or_404(
        QuotationRequest, pk=quotation_request_id)

    if request.method == 'POST':

        quotation_request.deadline_date = request.POST.get('deadline_date')

        if request.FILES.get('specification'):
            quotation_request.specification = request.FILES['specification']
        elif not quotation_request.specification:
            quotation_request.specification = quotation_request.specification

        if request.FILES.get('boq'):
            quotation_request.boq = request.FILES['boq']
        elif not quotation_request.boq:
            quotation_request.boq = quotation_request.boq

        quotation_request.save()
        messages.success(request, 'request edited successfully')
        return redirect('user_dashboard')

    context = {'quotation_request': quotation_request}
    return render(request, 'user/edit_request.html', context)

# view specific supplier bids(user)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def view_suppliers_bids(request, quotation_request_id):
    quotation_request = QuotationRequest.objects.get(id=quotation_request_id)
    supplier_bids = SupplierBid.objects.filter(
        quotation_request=quotation_request_id)
    unique_supplier_ids = supplier_bids.values_list(
        'user_id', flat=True).distinct()
    unique_supplier_bids = []
    for user_id in unique_supplier_ids:
        supplier_bid = supplier_bids.filter(user_id=user_id).first()
        unique_supplier_bids.append(supplier_bid)
    today = date.today()
    today = date.today()
    return render(request, 'user/supplier_bids.html', {'quotation_request': quotation_request, 'supplier_bids': supplier_bids, 'today': today, 'unique_supplier_bids': unique_supplier_bids})

# view supplier bids on line(user)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def view_bids_on_line(request, quotation_request_id, supplier_bid_id):

    bids = SupplierBid.objects.filter(
        quotation_request=quotation_request_id, user_id=supplier_bid_id)
    for bid in bids:
        bid.price_per_unit = round(bid.price / bid.quotation_line.quantity, 2)
    return render(request, 'user/bids_on_line.html', {'bids': bids})

# report view(user)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def report(request):
    bids = QuotationRequest.objects.filter(
        supplierbid__isnull=False, deadline_date__lte=date.today()).distinct()
    search_keyword = request.GET.get('search')
    if search_keyword:
        bids = bids.filter(quotation_number__icontains=search_keyword)

    paginator = Paginator(bids, 10)
    page = request.GET.get('page')
    try:
        bids = paginator.page(page)
    except PageNotAnInteger:
        bids = paginator.page(1)
    except EmptyPage:
        bids = paginator.page(paginator.num_pages)

    context = {'bids': bids}
    return render(request, 'user/report.html', context)

# view quotaion request (user)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def request_detail(request, quotation_id):
    lines = QuotationLine.objects.filter(
        quotation_request=quotation_id).order_by('line_number')
    if lines.exists():
        requisition_no = lines.first().quotation_request.quotation_number
    else:
        requisition_no = None
    if request.method == "GET":
        checked_lines = request.GET.get("checked_lines", "")
        checked_lines_list = [int(x) for x in checked_lines.split(",") if x]
        if checked_lines_list:
            return redirect('send')

    context = {'lines': lines,
               'requisition_no': requisition_no,
               'checked_lines_list': checked_lines_list}
    return render(request, 'user/quotation_detail.html', context)

# view to send request to suppliers(user)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def send_quotation(request):
    checked_lines = request.GET.getlist("checked_lines")
    requesition_number = request.GET.getlist("reqestion_number")
    email_data = request.POST.get("emailData")
    telegram_data = request.POST.get("telegramData")
    checked_lines_list = [int(num)
                          for num in request.GET.get('checked_lines').split(',')]
    requesition_number = int(request.GET.get('reqestion_number'))

    encoded_id = hashids.encode(requesition_number)
    encoded_lines = hashids.encode(*checked_lines_list)

    linki = f"{encoded_id}&checked_lines={encoded_lines}"

    sms_link1 = "http://127.0.0.1:8000/login/?id=" + linki
    sms_link2 = "http://127.0.0.1:8000/signup/?id=" + linki
    sms_message = "ውድ ደንበኞቻችን,\nአዋሽ ወይን ጠጅ እቃዎችን አወዳድሮ ለመግዛት በድረ-ገጽ ጨረታ የጀመረ በመሆኑ ሊንኩን በመጫን ዋጋ እንዲያስገቡ ተጋብዘዋል፡፡\n" + sms_link1 + "\nአካውንት ከሌሎ ከስር ያለውን ሊንክ በመጫን አካውንት ይክፈቱ፡፡\n" + sms_link2 + \
        "\nDear Supplier,\nAwash Wines has launched an online bid to compare suppliers, so click on the link to submit your bid price\n" + \
        sms_link1 + "\nIf you don't have an account you can register by clicking the link below\n" + sms_link2
    email_message = f"""
<!DOCTYPE html>
<html lang='en' xmlns='http://www.w3.org/1999/xhtml' xmlns:o='urn:schemas-microsoft-com:office:office'>
<head>
<meta charset='UTF-8'>
<meta name='viewport' content='width=device-width,initial-scale=1'>
<meta name='x-apple-disable-message-reformatting'>
<title></title>
<!--[if mso]>
<noscript>
 <xml>
   <o:OfficeDocumentSettings>
     <o:PixelsPerInch>96</o:PixelsPerInch>
   </o:OfficeDocumentSettings>
 </xml>
</noscript>
<![endif]-->
<style>
</style>
</head>
<body style='margin:0;padding:0;'>
<table role='presentation' style='width:100%;border-collapse:collapse;border:0;border-spacing:0;background:#ffffff;'>
 <tr>
 <td align='center' style='padding:0;'>
 <table role='presentation' style='width:602px;border-collapse:collapse;border:1px solid #cccccc;border-spacing:0;text-align:left;'>
   <tr>
     <td align='center' style='padding:40px 0 30px 0;background:#E9EDF4;'>
       <img src='http://awashwines.com/wp-content/uploads/AWASH-WINE-LOGO-1.png' />
       <hr style='margin: 30px;'>
     </td>
       </tr>
       <tr>
         <td style='padding:36px 30px 42px 30px;background:#E9EDF4;'>
           <table role='presentation' style='width:100%;border-collapse:collapse;border:0;border-spacing:0;'>
             <tr>
               <td style='padding:0 0 36px 0;color:#153643;background:#E9EDF4;'>
                 <h1 style='font-size:24px;margin:0 0 20px 0;font-family:Arial,sans-serif;'>Awash Wine Quotation Request</h1>
                 <p style='margin:0 0 12px 0;line-height:24px;font-family:Arial,sans-serif;'><b> Dear Supplier </b><br>
                 </p>
                 <p style='margin:0;line-height:24px;font-family:Arial,sans-serif;'>
                 Awash Wine has launched an online bid  and you can register by clicking the Sign Up button to participate in the bid.
                   </p>
                 <p><p/><a href='{sms_link2}'><button style='background-color: #3F84FC; border: none; border-radius: 5px; color: white; /* Dark grey */ padding: 15px 32px'>signup</button></a><br><br><br><br>
                 <p style='margin:0;line-height:24px;font-family:Arial,sans-serif;'>
                 We will be accepting bid proposals for different items you will find listed on the system.
                  If you are interested in submitting a bid, please click in Login Button.
                 </p>
                 <form action='' method='post'>
                  <a href='{sms_link1}'><button name='link' style='background-color: #3F84FC; border: none; border-radius: 5px; color: white; /* Dark grey */ padding: 15px 32px'>signin</button></a>
                  <p style='margin:0;line-height:24px;font-family:Arial,sans-serif;'>
                  Bids must be prepared and submitted as per the detail requested by the system . Any bid submitted through the system  would be considered  as original PFI signed and stamped by the biding company.
                Bids received after the bid closing date and time will not be accepted.
                All bids received by the above-mentioned deadline shall be reviewed and  winner will be communicated through  email/phone.<br>
                Regards<br>
                Awash wine Procurement Team
                  </p>
                  <p style='margin:0 0 12px 0;line-height:24px;font-family:amharic;'><b> ውድ ደንበኞቻችን  </b><br>አዋሽ ወይን ኦንላይን የጨረታ ስርዓት የጀመረ ሲሆን በጨረታው ለመሳተፍ sign up የሚለውን የመመዝገቢያ ቁልፍ በመጫን መመዝገብ ይችላሉ፡፡<br></p><a href='{sms_link1}'><button style='background-color: #3F84FC; border: none; border-radius: 5px; color: white; /* Dark grey */ padding: 15px 32px'>signup</button></a><br><br><br><br>
                  <p style='margin:0;line-height:24px;font-family:amharic;'>
                በኦንላይን መርሃ ግብሩ ላይ ተዘርዝረው የሚያገኙአቸው የተለያዩ እቃዎች ጨረታ በመቀበል ላይ ስለሆንን በጨረታው ለመሳተፍ ከፈለጉ sign in የሚለውን መግቢያ ቁልፍ ተጭነው መሳተፍ ይችላሉ፡፡
                  </p>
                  <form action='' method='post'>
                   <a href='{sms_link1}'><button name='link' style='background-color: #3F84FC; border: none; border-radius: 5px; color: white; /* Dark grey */ padding: 15px 32px'>signin</button></a>
                   <p style='margin:0;line-height:24px;font-family:amharic;'>
                   ተጫራቾች ኦንላይን መርሃ ግብሩ ላይ በተጠየቀው ዝርዝር መሰረት የዋጋ መረጃ ማቅረብ አለባቸው፡፡ በመርሃ ግብሩ ላይ የምታስገቡት ማንኛውም ዋጋ በድርጅታችሁ የተፈረመ እና ማህተም የተደረገ ኦርጂናል ዋጋ ማቅረቢያ ተደርጎ ይቆጠራል ::
                   ጨረታው ከሚዘጋበት ቀን እና ሰዓት በኋላ የገባ የጨረታ ማቅረቢያ ተቀባይነት የለውም፡፡ ጨረታውን ለማስገባት በተሰጠው ጊዜ ገደብ ውስጥ ካስገቡ ተጫራቾች መሃከል አሸናፊውን በኢሜል /በስልክ የምንገልፅ ይሆናል፡፡<br>
                ከሰላምታ ጋር <br>
                የአዋሽ ወይን ግዥ ክፍል

                   </p>
               </td>
             </tr>
</form>
           </table>
         </td>
       </tr>
       <tr>
       <td style='padding:30px;background:#E9EDF4;'><hr style='margin: 30px;'>
       <table role='presentation' style='width:100%;border-collapse:collapse;border:0;border-spacing:0;font-size:9px;font-family:Arial,sans-serif;'>
         <tr>
           <td style='padding:0;width:50%;' align='left'>
            <center> <p style='margin:0;font-size:14px;line-height:16px;font-family:Arial,sans-serif;color:#153643;'>
               &reg; Awash Wine S.C, Ethiopia 2024
             </p></center>
           </td>

             </tr>
           </table>
         </td>
       </tr>
     </table>
   </td>
 </tr>
</table>
</body>
</html>
   """

    if email_data:
        email_data = json.loads(email_data)
        for email in email_data:
            subject = "AWASH WINE S.C RFQ"

            from_email = "awash.rfq@gmail.com"
            email = email.strip()

            msg = EmailMessage(subject, email_message, from_email, [email])
            msg.content_subtype = 'html'
            msg.send()
        messages.success(request, f'Email sent successfully')

    if telegram_data:
        telegram_data = json.loads(telegram_data)
        for telegram in telegram_data:
            send_message('192.168.110.11', 9797, 'admin',
                         '@dIff%nSms0', telegram, sms_message)
        messages.success(request, f'SMS sent successfully')

    if telegram_data or email_data:
        return redirect('user_dashboard')

    context = {}
    return render(request, 'user/send.html', context)

# view signup


def signup(request):
    id_value = request.GET.get('id')
    checked_lines_value = request.GET.get('checked_lines')
    request.session['id'] = id_value
    request.session['checked_lines'] = checked_lines_value
    request.session.modified = True
    urls = f'?id={id_value}&checked_lines={checked_lines_value}'
    if request.method == 'POST':
        cname = request.POST.get('cname')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        password = request.POST.get('password')
        con_password = request.POST.get('con_password')
        if password == con_password:
            try:
                if User.objects.filter(username=email).exists():
                    messages.error(request, f'User already exist')
                    redirect_url = reverse(
                        'signup') + f'?id={id_value}&checked_lines={checked_lines_value}'
                    return redirect(redirect_url)
                if User.objects.filter(email=email).exists():
                    messages.error(request, f'User already exist')
                    redirect_url = reverse(
                        'signup') + f'?id={id_value}&checked_lines={checked_lines_value}'
                    return redirect(redirect_url)
                if User.objects.filter(last_name=phone).exists():
                    messages.error(request, f'User already exist')
                    redirect_url = reverse(
                        'signup') + f'?id={id_value}&checked_lines={checked_lines_value}'
                    return redirect(redirect_url)
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=password,
                    first_name=cname,
                    last_name=phone
                )
                group = Group.objects.get(name='supplier')
                user.groups.add(group)
                messages.success(request, f'Account Created Successfully')
                redirect_url = reverse(
                    'login') + f'?id={id_value}&checked_lines={checked_lines_value}'
                return redirect(redirect_url)
            except Exception as e:
                print(e)
                messages.error(request, f'Failed to create user')
                redirect_url = reverse(
                    'signup') + f'?id={id_value}&checked_lines={checked_lines_value}'
                return redirect(redirect_url)

        else:
            messages.error(request, f'Password do not match')
            redirect_url = reverse(
                'signup') + f'?id={id_value}&checked_lines={checked_lines_value}'
            return redirect(redirect_url)

    context = {'urls': urls}
    return render(request, 'signup.html', context)

# view supplier dashboard(supplier)


@login_required(login_url='login')
@user_passes_test(user_in_supplier_group, login_url='login')
def supplier_dashboard(request):
    id_value = request.session.get('id')
    currenrt_date = date.today()
    checked_lines_value = request.session.get('checked_lines')
    if id_value:
        decoded_id = hashids.decode(id_value)[0]
        decoded_line = hashids.decode(checked_lines_value)
        decoded_line = list(decoded_line)
        print(decoded_id)
        print(decoded_line)
        quotation_request = QuotationRequest.objects.filter(
            quotation_number=decoded_id)
        quotation_lines = QuotationLine.objects.filter(
            Q(quotation_request__in=quotation_request) &
            Q(line_number__in=decoded_line)
        ).order_by('line_number')

    context = {'quotation_lines': quotation_lines,
               'currenrt_date': currenrt_date}
    return render(request, 'supplier/index.html', context)

# view bid history(supplier)


@login_required(login_url='login')
@user_passes_test(user_in_supplier_group, login_url='login')
def bid_history(request):
    current_date = date.today()
    supplier_bids = SupplierBid.objects.filter(user_id=request.user)
    quotation_requests = QuotationRequest.objects.filter(
        pk__in=[bid.quotation_request.pk for bid in supplier_bids]
    ).order_by('deadline_date')
    search_keyword = request.GET.get('search')
    if search_keyword:
        quotation_requests = quotation_requests.filter(
            quotation_number__icontains=search_keyword)
    paginator = Paginator(quotation_requests, 10)  # Show 10 items per page
    page = request.GET.get('page')
    try:
        quotation_requests = paginator.page(page)
    except PageNotAnInteger:
        quotation_requests = paginator.page(1)
    except EmptyPage:
        quotation_requests = paginator.page(paginator.num_pages)
    context = {'quotation_requests': quotation_requests,
               'current_date': current_date}
    return render(request, 'supplier/bid_history.html', context)

# view bid(supplier)


@login_required(login_url='login')
@user_passes_test(user_in_supplier_group, login_url='login')
def bid(request):
    id_value = request.session.get('id')
    checked_lines_value = request.session.get('checked_lines')
    decoded_id = hashids.decode(id_value)[0]
    decoded_line = hashids.decode(checked_lines_value)
    decoded_line = list(decoded_line)
    quotation_request = QuotationRequest.objects.filter(
        quotation_number=decoded_id)
    quotation_lines = QuotationLine.objects.filter(
        Q(quotation_request__in=quotation_request) &
        Q(line_number__in=decoded_line)
    )

    if request.method == 'POST':
        vat_type = request.POST.get('vat_type')
        quotation_request_id = request.POST.getlist('quotation_request')[0]
        print(quotation_request_id)
        quotation_request = QuotationRequest.objects.get(
            id=quotation_request_id)

        if quotation_request.deadline_date.date() <= date.today():
            messages.error(
                request, "This request is already closed. You cannot place a bid.")

        else:
            payment = request.POST.get('payment')
            lead_time = request.POST.get('lead_time')
            delivery_term = request.POST.get('delivery_term')
            validity_date = request.POST.get('validity_date')
            boq_file = request.FILES.get('boq')
            attach_file = request.FILES.get('attach')

            bid_exists = False
            for decoded_lines in decoded_line:
                price = request.POST.get(f'total_price{decoded_lines}')
                print(price)
                line_number = QuotationLine.objects.get(
                    line_number=decoded_lines, quotation_request=quotation_request)
                try:
                    bid = SupplierBid.objects.get(
                        quotation_request=quotation_request,
                        user_id=request.user,
                        quotation_line=line_number
                    )
                    if not bid_exists:
                        messages.error(request, 'Bid already exists')
                        bid_exists = True

                except SupplierBid.DoesNotExist:
                    bid = SupplierBid(
                        quotation_request=quotation_request,
                        quotation_line=line_number,
                        user_id=request.user,
                        price=price,
                        vat=vat_type,
                        delivery_term=delivery_term,
                        validity_date=validity_date,
                        payment=payment,
                        lead_time=lead_time,
                        file=attach_file,
                        boq_total=boq_file,
                    )
                    bid.save()
                    if not bid_exists:
                        messages.success(
                            request, 'Bid submitted successfully!')
                        bid_exists = True

    return render(request, './supplier/bid.html', context={'quotation_lines': quotation_lines})


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def user_profile(request):
    user = request.user
    if request.method == 'POST':
        user = request.user
        old_password = request.POST.get('password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        if not user.check_password(old_password):
            messages.error(request, 'The old password is incorrect.')
            return redirect('user_profile')
        if new_password != confirm_password:
            messages.error(request, 'The new passwords do not match.')
            return redirect('user_profile')
        user.set_password(new_password)
        user.save()
        update_session_auth_hash(request, user)

        messages.success(
            request, 'Your password has been updated successfully.')

        return redirect('user_profile')
    context = {
        'user': user,
    }
    return render(request, './user/profile.html', context)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def supplier_profile(request):
    user = request.user  # Assuming you're using Django's built-in User model
    if request.method == 'POST':
        user = request.user
        old_password = request.POST.get('password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        # Check if the old password is correct
        if not user.check_password(old_password):
            messages.error(request, 'The old password is incorrect.')
            # Redirect to profile page if old password is incorrect
            return redirect('supplier_profile')

        # Check if new password and confirm password match
        if new_password != confirm_password:
            messages.error(request, 'The new passwords do not match.')
            # Redirect to profile page if passwords don't match
            return redirect('supplier_profile')

        # Set the new password and update session authentication hash
        user.set_password(new_password)
        user.save()
        update_session_auth_hash(request, user)  # Important to update session

        messages.success(
            request, 'Your password has been updated successfully.')
        # Redirect to profile page after successful update
        return redirect('supplier_profile')
    context = {
        'user': user,
    }
    return render(request, './supplier/profile.html', context)


# view for supplier base
@login_required(login_url='login')
@user_passes_test(user_in_supplier_group, login_url='login')
def supplier_base(request):
    user = request.user
    context = {
        'user': user,
    }
    return render(request, 'supplier_main.html', context)

# view for user base


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def user_base(request):
    user = request.user
    context = {
        'user': user,
    }
    return render(request, 'user_main.html', context)

# view bid history on line(supplier)


@login_required(login_url='login')
@user_passes_test(user_in_supplier_group, login_url='login')
def view_bid_history(request, quotation_line_id):

    suppliers_bids = SupplierBid.objects.filter(
        user_id=request.user,
        quotation_request_id=quotation_line_id
    )
    for bid in suppliers_bids:
        bid.price_per_unit = round(bid.price / bid.quotation_line.quantity, 2)
    user = request.user
    context = {
        'user': user,
        'suppliers_bids': suppliers_bids
    }
    return render(request, './supplier/view_bid_history.html', context)

# download bid excell(user)


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def download_bids_excel(request, quotation_request_id, user_id):
    bids = SupplierBid.objects.filter(
        quotation_request=quotation_request_id, user_id=user_id)
    wb = Workbook()
    ws = wb.active
    user = User.objects.get(id=user_id)
    ws['A1'] = 'Line Number'
    ws['B1'] = 'Description'
    ws['C1'] = 'Quantity'
    ws['D1'] = 'Unit Price'
    ws['E1'] = 'Total Price'
    ws['F1'] = 'Grand Total'
    ws['G1'] = 'Lead Time'
    ws['H1'] = 'Created'

    for bid in bids:
        created_at_date = bid.created_at.date()
        if bid.vat == 1:
            vat = 0.15
        else:
            vat = 0.02
        created_at_str = created_at_date.strftime("%Y, %B, %d")
        ws.append([
            bid.quotation_line.line_number,
            bid.quotation_line.item_description,
            bid.quotation_line.quantity,
            round(bid.price/bid.quotation_line.quantity, 2),
            bid.price,
            round(float(bid.price)*float(vat)+float(bid.price), 2),
            bid.lead_time,
            created_at_str
        ])
    file_name = f'bids_for_quotation_request_{
        quotation_request_id}_user_{user.first_name}.xlsx'
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type=content_type, headers={
        'Content-Disposition': f'attachment; filename="{file_name}"'
    })

# view for generating comparision


@login_required(login_url='login')
@user_passes_test(user_in_users_group, login_url='login')
def generate_price_comparison_excel(request, quotation_request_id):
    quotation_request = get_object_or_404(
        QuotationRequest, id=quotation_request_id)
    quotation_lines = QuotationLine.objects.filter(
        quotation_request=quotation_request)
    suppliers = SupplierBid.objects.filter(
        quotation_line__in=quotation_lines).values_list('user_id', flat=True).distinct()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Price Comparison'
    column_widths = {'A': 40, 'B': 10}
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color='BFBFBF',
                              end_color='BFBFBF', fill_type='solid')
    grey_fill = PatternFill(
        start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')

    ws.merge_cells('A1:C1')
    ws['A1'].value = f'PR no-{quotation_request.quotation_number} '
    ws['A1'].font = Font(bold=True, size=18)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = 'Awash Wine Price Comparison'
    ws['A2'].font = Font(bold=True, size=16)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:C3')
    ws['A3'] = 'Bid Information'
    ws['A3'].font = Font(bold=True, size=14)
    ws['A5'] = 'Description'
    ws['A5'].font = header_font
    ws['A5'].fill = header_fill
    ws['B5'] = 'Qty'
    ws['B5'].font = header_font
    ws['B5'].fill = header_fill
    supplier_columns = {}
    for idx, supplier_id in enumerate(suppliers, start=1):
        supplier_name = User.objects.get(id=supplier_id).first_name
        ws.cell(row=5, column=idx + 2, value=supplier_name)
        ws.cell(row=5, column=idx + 2).font = header_font
        ws.cell(row=5, column=idx + 2).fill = header_fill
        supplier_columns[supplier_id] = idx + 2
    row_num = 6
    for quotation_line in quotation_lines:
        ws.cell(row=row_num, column=1, value=quotation_line.item_description)
        ws.cell(row=row_num, column=1).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_num, column=2, value=quotation_line.quantity)
        for supplier_id in suppliers:
            bid = SupplierBid.objects.filter(
                quotation_line=quotation_line, user_id=supplier_id).first()
            if bid:
                calculated_price = float(bid.price)
                vat_multiplier = 0.15 if bid.vat == 1 else 0.02
                total_price_with_vat = calculated_price * (1 + vat_multiplier)
                ws.cell(row=row_num,
                        column=supplier_columns[supplier_id], value=bid.price)
        row_num += 1
    total_price_row = row_num
    ws.cell(row=total_price_row, column=1, value='Total Price')
    for supplier_id in suppliers:
        total_price = sum(float(bid.price) for bid in SupplierBid.objects.filter(
            quotation_line__in=quotation_lines, user_id=supplier_id))
        ws.cell(row=total_price_row,
                column=supplier_columns[supplier_id], value=total_price)
    for cell in ws[total_price_row]:
        cell.fill = grey_fill
    row_num += 1
    grand_total_row = row_num
    ws.cell(row=grand_total_row, column=1, value='Grand Total (incl. VAT)')
    for supplier_id in suppliers:
        total_price = sum(float(bid.price) for bid in SupplierBid.objects.filter(
            quotation_line__in=quotation_lines, user_id=supplier_id))
        total_vat_multiplier = 0.15 if bid.vat == 1 else 0.02
        grand_total_with_vat = total_price * (1 + total_vat_multiplier)
        ws.cell(row=grand_total_row,
                column=supplier_columns[supplier_id], value=grand_total_with_vat)
    for cell in ws[grand_total_row]:
        cell.fill = grey_fill
    total_price_values = {supplier_id: sum(float(bid.price) for bid in SupplierBid.objects.filter(
        quotation_line__in=quotation_lines, user_id=supplier_id)) for supplier_id in suppliers}
    sorted_suppliers = sorted(total_price_values.items(), key=lambda x: x[1])
    if sorted_suppliers:
        winner_supplier_id, _ = sorted_suppliers[0]
        if winner_supplier_id in supplier_columns:
            winner_column_index = supplier_columns[winner_supplier_id] - 1
            for row in ws.iter_rows(min_row=6, max_row=grand_total_row, min_col=winner_column_index + 1, max_col=winner_column_index + 1):
                for cell in row:
                    cell.fill = PatternFill(
                        start_color='add8e6', end_color='add8e6', fill_type='solid')
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(
    ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="price_comparison_{
        quotation_request_id}.xlsx"'
    return response
